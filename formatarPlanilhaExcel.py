import openpyxl
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import NamedStyle
import pandas as pd
import locale
import os

# Configurar o locale para português do Brasil
locale.setlocale(locale.LC_TIME, 'pt_BR.UTF-8')

from logger import infoLogs
from caminhos import EMPRESAS,caminhoPlanilhaNotasDebito,caminho_planilha_notas_CAP_aberto,planilhaPastaNotasCanceladas,caminho_planilha_notas_CAP,CAMINHO_DESTINO, CAMINHO_NOTAS_DEBITOS,CAMINHO_NOTAS_CREDITOS,CAMINHO_NOTAS_PAGAMENTOS,CAMINHO_NOTAS_CANCELADAS
from datas import carregar_datas






# Mapeamento manual para corrigir problemas de codificação
meses_em_portugues = {
    "janeiro": "Janeiro",
    "fevereiro": "Fevereiro",
    "março": "Março",
    "abril": "Abril",
    "maio": "Maio",
    "junho": "Junho",
    "julho": "Julho",
    "agosto": "Agosto",
    "setembro": "Setembro",
    "outubro": "Outubro",
    "novembro": "Novembro",
    "dezembro": "Dezembro",
}



def limpar_boletos_e_excluir_cabeçalhos(empresa):


    
    arquivo_destino = rf'{CAMINHO_DESTINO}\planilha_principal_{empresa}.xlsx'  # Arquivo onde as abas serão criadas

    



    infoLogs().info(f"ETAPA LIMPAR BOLETOS E EXCLUIR CABEÇALHOS - INICIADA | {empresa}")

    try:
        # Carregar a planilha
        wb = load_workbook(arquivo_destino)

        # Textos a serem localizados e cujas linhas serão limpas
        textos_para_limpar = ['Boletos Abertos', 'Boletos Pagos', 'Total']

        # Função para excluir linhas que contenham 'VENCTO' ou estejam vazias
        def excluir_linhas(ws):
            max_row = ws.max_row
            
            # Percorrer todas as linhas de baixo para cima para evitar problemas ao excluir
            for row in range(max_row, 1, -1):
                # Verificar se a célula da coluna "VENCTO" (primeira coluna) é 'VENCTO'
                if ws.cell(row=row, column=1).value == "VENCTO":
                    ws.delete_rows(row)
                # Verificar se a linha inteira está vazia
                elif all(cell.value is None for cell in ws[row]):
                    ws.delete_rows(row)

        # Percorrer todas as abas da planilha
        for ws in wb.worksheets:
            # infoLogs().info(f"Processando a aba {ws.title}")

            # Limpar as linhas de boletos
            for row in range(1, ws.max_row + 1):
                cell_value = ws.cell(row=row, column=6).value  # Coluna F (6) com o texto a ser verificado
                if cell_value in textos_para_limpar:
                    # Limpar o conteúdo da linha inteira
                    for col in range(1, ws.max_column + 1):
                        ws.cell(row=row, column=col).value = None

            # Excluir cabeçalhos e linhas vazias
            excluir_linhas(ws)

        # Salvar a planilha com as alterações
        wb.save(arquivo_destino)
        infoLogs().info(f"ETAPA LIMPAR BOLETOS E EXCLUIR CABEÇALHOS - FINALIZADA | {empresa}")
    except Exception as e:
        infoLogs().info(f"Erro ao processar a planilha: {e} | {empresa}")

    infoLogs().info(f"Boletos e cabeçalhos foram processados com sucesso. | {empresa}")

    
    

def copia_notas_debitos(empresa):
    
    try:
        # Caminhos dos arquivos
        # arquivo_origem = rf'{CAMINHO_NOTAS_DEBITOS}\notas_debitos_05-12-2024.xlsx'  # Arquivo com os dados de origem
        arquivo_origem = caminhoPlanilhaNotasDebito(empresa)

        arquivo_destino = rf'{CAMINHO_DESTINO}\planilha_principal_{empresa}.xlsx'

        


        # Carregar as planilhas
        wb_origem = openpyxl.load_workbook(arquivo_origem)
        wb_destino = openpyxl.load_workbook(arquivo_destino)
        planilha_origem = wb_origem.active  # Planilha de origem

        # Cabeçalhos fixos e colunas desejadas
        cabecalhos = ["VENCTO", "RECBO", "EMISSAO", "PGs", "NFISCAL", "NOME", "CREDITO", "DEBITO", "OBS"]
        colunas_desejadas = {
            "VENCTO": 1,  # Índice da coluna na planilha de origem
            "RECBO": 2,
            "EMISSAO": 3,
            "PGs": 4,
            "NFISCAL": 5,
            "NOME": 6,
            "CREDITO": 7,
        
            # Adicionar colunas adicionais conforme necessário
        }

        # Função para converter valores em datetime
        def converter_para_data(valor):
            if isinstance(valor, datetime):
                return valor
            try:
                return datetime.strptime(valor, "%d/%m/%Y")
            except (ValueError, TypeError):
                return None

        # Processar as linhas da planilha de origem
        for linha in planilha_origem.iter_rows(min_row=2, values_only=True):  # Ignorar cabeçalho
            data = converter_para_data(linha[0])  # Supondo que a data está na coluna A
            if not data:
                continue  # Ignorar linhas com datas inválidas ou vazias

            # Criar nome da aba no formato Mês-Ano
            nome_aba = data.strftime("%B-%Y").capitalize()

            try:
                # Substituir "Marã§o" por "Marco" (garantindo que o nome correto será usado)
                nome_aba = nome_aba.replace("Marã§o", "Marco")
            except Exception as e:
                infoLogs().info(f'Nao existe o valor de Marã§o notas_creditos_aberto - {e}')

            # Criar a aba, se não existir
            if nome_aba not in wb_destino.sheetnames:
                aba_destino = wb_destino.create_sheet(title=nome_aba)
                # Adicionar os cabeçalhos
                for col_idx, cabecalho in enumerate(cabecalhos, start=1):
                    aba_destino.cell(row=1, column=col_idx, value=cabecalho)
            else:
                aba_destino = wb_destino[nome_aba]

            # Encontrar a próxima linha vazia na aba de destino
            for linha_idx in range(2, aba_destino.max_row + 2):  # Começar após os cabeçalhos
                if all(aba_destino.cell(row=linha_idx, column=col_idx).value is None for col_idx in range(1, len(cabecalhos) + 1)):
                    linha_destino = linha_idx
                    break

            # Copiar os valores das colunas desejadas
            for col_destino, col_origem in enumerate(colunas_desejadas.values(), start=1):
                valor = linha[col_origem - 1]  # Ajustar para índice baseado em 0
                if aba_destino.cell(row=linha_destino, column=col_destino).value is None:  # Não sobrescrever valores existentes
                    aba_destino.cell(row=linha_destino, column=col_destino, value=valor)

        # Salvar o arquivo de destino
        wb_destino.save(arquivo_destino)
        infoLogs().info(f"Processo concluído! Dados copiados de notas_debitos para planilha_principal_lokan | {empresa}")
    except Exception as e:
        infoLogs().info(f"Não existem planilha notas_debitos da empresa {empresa} para o periodo informado - \n{e}")   


    
    






def copia_notas_creditos_em_aberto(empresa):
    try:

        arquivo_origem = caminho_planilha_notas_CAP_aberto(empresa)

        arquivo_destino = rf'{CAMINHO_DESTINO}\planilha_principal_{empresa}.xlsx'

        

        # Carregar as planilhas
        wb_origem = openpyxl.load_workbook(arquivo_origem)
        wb_destino = openpyxl.load_workbook(arquivo_destino)
        planilha_origem = wb_origem.active  # Planilha de origem

        # Cabeçalhos fixos
        cabecalhos_destino = ["VENCTO", "RECBO", "EMISSAO", "PGs", "NFISCAL", "NOME", "CREDITO", "DEBITO", "OBS"]
        
        colunas_desejadas = {
            "VENCTO": 1,  # Índice da coluna na planilha de origem
            "RECBO": 2,
            "EMISSAO": 3,
            "PGs": 4,
            "NFISCAL": 5,
            "NOME": 6,
            "CREDITO": 7,
            "DEBITO": 8,
            "OBS": 9,
        }

    
        # Função para converter valores em datetime
        def converter_para_data(valor):
            if isinstance(valor, datetime):
                return valor
            for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
                try:
                    return datetime.strptime(str(valor), fmt)
                except (ValueError, TypeError):
                    continue
            return None

        # Processar as linhas da planilha de origem
        for linha in planilha_origem.iter_rows(min_row=2, values_only=True):  # Ignorar cabeçalho
            data = converter_para_data(linha[0])  # Supondo que a data está na coluna A
            if not data:
                continue  # Ignorar linhas com datas inválidas ou vazias

            # Criar nome da aba no formato Mês-Ano
            nome_aba = data.strftime("%B-%Y").capitalize()

            try:
                # Substituir "Marã§o" por "Marco" (garantindo que o nome correto será usado)
                nome_aba = nome_aba.replace("Marã§o", "Marco")
            except Exception as e:
                infoLogs().info(f'Nao existe o valor de Marã§o notas_creditos_aberto - {e}')

            # Criar a aba, se não existir
            if nome_aba not in wb_destino.sheetnames:
                aba_destino = wb_destino.create_sheet(title=nome_aba)
                # Adicionar os cabeçalhos
                for col_idx, cabecalho in enumerate(cabecalhos_destino, start=1):
                    aba_destino.cell(row=1, column=col_idx, value=cabecalho)
            else:
                aba_destino = wb_destino[nome_aba]

            # Mapear os cabeçalhos das colunas de destino
            cabecalhos_aba = {
                aba_destino.cell(row=1, column=col_idx).value: col_idx
                for col_idx in range(1, aba_destino.max_column + 1)
            }

            # Encontrar a próxima linha vazia na aba de destino
            linha_destino = aba_destino.max_row + 1

            # Copiar os valores para as colunas corretas
            for cabecalho, col_origem in colunas_desejadas.items():
                valor = linha[col_origem - 1]  # Ajustar para índice baseado em 0
                if cabecalho in cabecalhos_aba:  # Verificar se o cabeçalho existe na aba
                    col_destino = cabecalhos_aba[cabecalho]
                    aba_destino.cell(row=linha_destino, column=col_destino, value=valor)

        # Salvar o arquivo de destino
        wb_destino.save(arquivo_destino)
        infoLogs().info(f"Processo concluído! Dados copiados de notas_creditos_em_aberto para {empresa}.")
    except Exception as e:
        infoLogs().info(f"Não existem planilha notas_creditos_em_aberto para empresa {empresa} no periodo informado - \n{e}")




    






def copia_notas_canceladas(empresa):
    try:
        arquivo_origem = planilhaPastaNotasCanceladas(empresa)

        arquivo_destino = rf'{CAMINHO_DESTINO}\planilha_principal_{empresa}.xlsx'

        # Carregar a planilha de destino
        wb_destino = openpyxl.load_workbook(arquivo_destino)

        # Obter a data atual e traduzir o mês manualmente
        data_atual = datetime.now()
        mes_em_portugues = meses_em_portugues[data_atual.strftime("%B").lower()]
        nome_aba_mes = f"{mes_em_portugues}-{data_atual.year}"

        try:
            # Substituir qualquer ocorrência errada do mês "Março" (Marã§o)
            nome_aba_mes = nome_aba_mes.replace("Marã§o", "Marco")
        except Exception as e:
            infoLogs().info(f'Nao existe o nome da aba Marã§o para notas_canceladas ')   

        # Verificar se a aba do mês atual existe, criar se não existir
        if nome_aba_mes not in wb_destino.sheetnames:
            ws_destino = wb_destino.create_sheet(title=nome_aba_mes)
            cabecalhos_destino = ["VENCTO", "RECBO", "EMISSAO", "PGs", "NFISCAL", "NOME", "CREDITO", "DEBITO", "OBS"]
            for col_idx, cabecalho in enumerate(cabecalhos_destino, start=1):
                ws_destino.cell(row=1, column=col_idx, value=cabecalho)
        else:
            ws_destino = wb_destino[nome_aba_mes]

        # Carregar a planilha de notas canceladas
        wb_canceladas = openpyxl.load_workbook(arquivo_origem)
        ws_canceladas = wb_canceladas.active

        # Função para encontrar a primeira linha vazia na planilha destino
        def find_first_empty_row(ws):
            for row in range(2, ws.max_row + 2):  # Começa após o cabeçalho
                if all(ws.cell(row=row, column=col).value is None for col in range(1, ws.max_column + 1)):
                    return row
            return ws.max_row + 1

        # Obter o dia 15 do mês atual
        dia_15 = data_atual.replace(day=15)

        # Copiar os valores da planilha canceladas, filtrando as datas de VENCTO
        rows_to_insert = []  # Linhas a serem inseridas

        for row in range(2, ws_canceladas.max_row + 1):  # Começa da linha 2
            vencto_value = ws_canceladas.cell(row=row, column=2).value  # Cancelamento está na coluna 2

            try:
                # Converter para uma data, caso seja uma string
                if isinstance(vencto_value, str):
                    vencto_value = datetime.strptime(vencto_value, "%d/%m/%Y")

                if isinstance(vencto_value, datetime) and vencto_value.month == data_atual.month:
                    rows_to_insert.append(row)  # Linha com VENCTO no mês vigente
            except ValueError:
                continue  # Ignorar valores que não podem ser convertidos para data

        # Inserir os valores da planilha de canceladas na planilha do mês vigente
        first_empty_row = find_first_empty_row(ws_destino)
        
        for row in rows_to_insert:
            ws_destino.cell(row=first_empty_row, column=1, value=ws_canceladas.cell(row=row, column=2).value)  # VENCTO = Cancelamento
            ws_destino.cell(row=first_empty_row, column=2, value=ws_canceladas.cell(row=row, column=2).value)  # RECBO = Cancelamento
            ws_destino.cell(row=first_empty_row, column=3, value=ws_canceladas.cell(row=row, column=1).value)  # EMISSAO = EMISSAO
            ws_destino.cell(row=first_empty_row, column=4, value="BX")  # PGs = BX
            ws_destino.cell(row=first_empty_row, column=5, value=ws_canceladas.cell(row=row, column=3).value)  # NFISCAL
            ws_destino.cell(row=first_empty_row, column=6, value=ws_canceladas.cell(row=row, column=4).value)  # NOME
            ws_destino.cell(row=first_empty_row, column=7, value='')  # CREDITO = 0
            ws_destino.cell(row=first_empty_row, column=8, value='')  # DEBITO = 0
            ws_destino.cell(row=first_empty_row, column=9, value=ws_canceladas.cell(row=row, column=5).value)  # OBS = Motivo
            first_empty_row += 1

        # Salvar a planilha de destino
        wb_destino.save(arquivo_destino)

        infoLogs().info(f"Processo concluído! Dados copiados de notas_canceladas para {empresa}.")
    except Exception as e:
        infoLogs().info(f"Não existem planilha notas_canceladas para empresa {empresa} no periodo informado - \n{e}")





def separar_datas(empresa):
    
    try:
        arquivo_destino = rf'{CAMINHO_DESTINO}\planilha_principal_{empresa}.xlsx'

        # Carregar o arquivo de destino
        wb = openpyxl.load_workbook(arquivo_destino)

        # Função para converter valores em datetime
        def converter_para_data(valor):
            if isinstance(valor, datetime):
                return valor
            try:
                return datetime.strptime(valor, "%d/%m/%Y")
            except (ValueError, TypeError):
                return None

        # Percorrer todas as abas da planilha
        for aba in wb.sheetnames:
            ws = wb[aba]

            # Obter os cabeçalhos da primeira linha
            cabecalhos = [cell.value for cell in ws[1]]

            # Criar listas para armazenar as linhas abaixo e acima do dia 15
            linhas_abaixo_dia_15 = []
            linhas_acima_dia_15 = []

            # Verificar as datas da coluna VENCTO (assumindo que a coluna A seja VENCTO)
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
                data = converter_para_data(row[0])  # Primeira coluna contém a data
                if data:
                    if data.day <= 15:
                        linhas_abaixo_dia_15.append(row)
                    else:
                        linhas_acima_dia_15.append(row)

            # Limpar a aba, mantendo apenas o cabeçalho
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    cell.value = None

            # Inserir as linhas abaixo do dia 15 logo após o cabeçalho
            linha_atual = 2
            for linha in linhas_abaixo_dia_15:
                for col_idx, valor in enumerate(linha, start=1):
                    ws.cell(row=linha_atual, column=col_idx, value=valor)
                linha_atual += 1

            # Adicionar 9 linhas vazias e o cabeçalho para as datas acima do dia 15
            linha_atual += 9
            for col_idx, cabecalho in enumerate(cabecalhos, start=1):
                ws.cell(row=linha_atual, column=col_idx, value=cabecalho)
            linha_atual += 1

            # Inserir as linhas acima do dia 15
            for linha in linhas_acima_dia_15:
                for col_idx, valor in enumerate(linha, start=1):
                    ws.cell(row=linha_atual, column=col_idx, value=valor)
                linha_atual += 1

        # Salvar o arquivo com as alterações
        wb.save(arquivo_destino)

        infoLogs().info(f"Processo concluído! As datas foram separadas em todas as abas | {empresa}")
    except Exception as e:
        infoLogs(f"Erro ao separar planilhas por 1º quinzena e 2º quinzena para empresa {empresa}\n {e}")





def formatar_colunas_credito_debito(caminho_arquivo,empresa):
    try:
        # Abrir o arquivo Excel
        wb = load_workbook(caminho_arquivo)
        
        # Iterar por todas as abas
        for ws in wb.worksheets:
            # Identificar as colunas CREDITO (G) e DEBITO (H)
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                # Verificar as colunas G (CREDITO) e H (DEBITO) diretamente
                credito_cell = row[6]  # Coluna G (CREDITO)
                debito_cell = row[7]   # Coluna H (DEBITO)

                # Aplicar a cor azul na fonte da coluna G (CREDITO), ignorando o valor "CREDITO"
                if credito_cell.value is not None and str(credito_cell.value).strip().upper() != "CREDITO":
                    credito_cell.font = Font(color="0000FF", name="Arial", size=12,bold=True)

                # Aplicar a cor vermelha na fonte da coluna H (DEBITO), ignorando o valor "DEBITO"
                if debito_cell.value is not None and str(debito_cell.value).strip().upper() != "DEBITO":
                    debito_cell.font = Font(color="FF0000", name="Arial", size=12,bold=True)

            # Formatar as colunas G e H como moeda brasileira
            for row in range(2, ws.max_row + 1):  # Ignorar a linha de cabeçalho
                # Ignorar valores "CREDITO" e "DEBITO" ao formatar como moeda
                credito_valor = ws[f'G{row}'].value
                debito_valor = ws[f'H{row}'].value

                if credito_valor is not None and str(credito_valor).strip().upper() != "CREDITO":
                    ws[f'G{row}'].number_format = 'R$ #,##0.00'
                if debito_valor is not None and str(debito_valor).strip().upper() != "DEBITO":
                    ws[f'H{row}'].number_format = 'R$ #,##0.00'

        

        # Salvar as alterações no arquivo
        wb.save(caminho_arquivo)
        infoLogs().info(f"Formatação aplicada com sucesso às colunas G e H | {empresa}")

    except Exception as e:
        infoLogs().info(f"Erro ao formatar colunas G e H para empresa {empresa}: {e}")




def aplicar_bordas_em_todas_as_abas(arquivo,empresa):
    """
    Aplica bordas em todas as células preenchidas de todas as abas de uma planilha.

    Args:
        arquivo (str): Caminho do arquivo Excel.
    """
    try:
        wb = load_workbook(arquivo)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        # Processar todas as abas
        for ws in wb.worksheets:
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=9):  # Limita até a coluna I (9)
                for cell in row:
                    cell.border = thin_border  # Aplica borda

        wb.save(arquivo)
        infoLogs().info(f"Bordas aplicadas com sucesso em todas as abas | {empresa}")
    except Exception as e:
        infoLogs().info(f"Erro ao aplicar bordas para empresa {empresa}: {e}")



def formatacaoFinal(empresa):
    infoLogs().info("ETAPA FORMATAÇÃO FINAL - INICIADA")
    try:

        arquivo_destino = rf'{CAMINHO_DESTINO}\planilha_principal_{empresa}.xlsx'

        # Carregar a planilha existente
        wb = load_workbook(arquivo_destino)
        
        # Iterar por todas as abas na pasta
        for ws in wb.worksheets:
            # infoLogs().info(f"Formatando aba: {ws.title}")
            
            # Encontrar a primeira linha vazia
            first_empty_row = None
            for row in range(1, ws.max_row + 1):
                if all(cell.value is None for cell in ws[row]):
                    first_empty_row = row
                    break
            if first_empty_row is None:
                first_empty_row = ws.max_row + 1

            # Inserir duas linhas vazias
            insert_row = first_empty_row + 2

            # Adicionar os textos 'Boletos Abertos', 'Boletos Pagos' e 'Total' na coluna 'PRODUTO'
            ws[f'F{insert_row - 1 }'] = 'Boletos Abertos'
            ws[f'F{insert_row }'] = 'Boletos Pagos'
            ws[f'F{insert_row + 1}'] = 'Total'

            # Configurar o início e o fim das linhas
            start_row = 2  # A partir da linha 2
            end_row = ws.max_row  # Última linha com dados

            # Encontrar a primeira linha totalmente vazia
            first_empty_row = None
            for row in range(2, ws.max_row + 1):  # Começar na linha 2
                if all(ws[f'{col}{row}'].value is None for col in 'ABCDEFGH'):  # Verificar todas as colunas relevantes
                    first_empty_row = row
                    break

            # Se não encontrar uma linha totalmente vazia, considerar que deve ir até a última linha de dados
            if first_empty_row is None:
                first_empty_row = ws.max_row + 1

            # Aplicar as fórmulas na linha de inserção
            insert_row = first_empty_row
            ws[f'G{insert_row + 1}'] = f'=SUMIF(D{start_row}:D{first_empty_row - 1},"AB",G{start_row}:G{first_empty_row - 1})'
            ws[f'G{insert_row + 2}'] = f'=SUMIF(D{start_row}:D{first_empty_row - 1},"PG",G{start_row}:G{first_empty_row - 1})'
            ws[f'G{insert_row + 3}'] = f'=G{insert_row + 1} + G{insert_row + 2}'

            ws[f'H{insert_row + 1}'] = f'=SUMIF(D{start_row}:D{first_empty_row - 1},"AB",H{start_row}:H{first_empty_row - 1})'
            ws[f'H{insert_row + 2}'] = f'=SUMIF(D{start_row}:D{first_empty_row - 1},"PG",H{start_row}:H{first_empty_row - 1})'
            ws[f'H{insert_row + 3}'] = f'=H{insert_row + 1} + H{insert_row + 2}'

    

                # Encontrar a última linha vazia
            last_row = ws.max_row
            while last_row > 0 and all(cell.value is None for cell in ws[last_row]):
                last_row -= 1

            # Pular duas linhas após a última linha vazia
            insert_row = last_row + 2  # Pula duas linhas após a última vazia

            # Verificar se os textos já existem e substituí-los, se necessário
            for offset, text in enumerate(['Boletos Abertos', 'Boletos Pagos', 'Total']):
                cell = ws[f'F{insert_row + offset}']
                if cell.value == text:
                    # Se já existir, substitua o texto
                    cell.value = text
                else:
                    # Caso contrário, apenas adiciona o texto
                    cell.value = text

            # Definir o intervalo de linhas para a fórmula (a partir da segunda linha até a última linha com dados)
            start_row = 2  # A partir da linha 2
            end_row = last_row  # Última linha preenchida com dados

            # Encontrar a segunda ocorrência de "CREDITO" na planilha (supondo que está na coluna "I")
            credit_count = 0
            credit_row = None
            for row in range(1, ws.max_row + 1):
                if ws[f'G{row}'].value == 'CREDITO':
                    credit_count += 1
                    if credit_count == 2:  # Encontrar a segunda ocorrência de "CREDITO"
                        credit_row = row
                        break

            if credit_row:
                # Inserir a fórmula ao lado de 'boleto A' na linha após o segundo "CREDITO"
                ws[f'G{insert_row}'] = f'=SUMIF(D{credit_row + 1}:D{end_row},"AB",G{credit_row + 1}:G{end_row})'
                ws[f'G{insert_row + 1}'] = f'=SUMIF(D{credit_row + 1}:D{end_row},"PG",G{credit_row + 1}:G{end_row})'
                ws[f'G{insert_row + 2}'] = f'=G{last_row + 2} + G{last_row + 3}'


                # Inserir a fórmula na coluna DEBITO da 2º quinzena
                ws[f'H{insert_row}'] = f'=SUMIF(D{credit_row + 1}:D{end_row},"AB",H{credit_row + 1}:H{end_row})'
                ws[f'H{insert_row + 1}'] = f'=SUMIF(D{credit_row + 1}:D{end_row},"PG",H{credit_row + 1}:H{end_row})'
                ws[f'H{insert_row + 2}'] = f'=H{last_row + 2} + H{last_row + 3}'



                ## Definir o intervalo de linhas a partir da linha de dados
                start_row = 2  # Supondo que a linha de cabeçalho é a linha 1
            
            # Formatação da fonte Arial 12 para todas as células, exceto a linha de cabeçalho
            arial_12_font = Font(name="Arial", size=12)
            for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row):
                for cell in row:
                    cell.font = arial_12_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            # Formatar as colunas 'G' e 'H' como moeda brasileira
            for row in range(2, ws.max_row + 1):
                ws[f'G{row}'].number_format = 'R$ #,##0.00'
                ws[f'H{row}'].number_format = 'R$ #,##0.00'

                    # Aplicar a formatação ao cabeçalho (linhas de cabeçalho que contêm os textos mencionados)
            header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
            header_font = Font(bold=True, name="Arial", size=12)

            # Lista de textos que você deseja identificar
            header_texts = ["VENCTO", "RECBO", "EMISSAO", "PGs", "NFISCAL", "NOME", "CREDITO", "DEBITO", "OBS"]

            # Iterar por todas as células da planilha
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value in header_texts:
                        # Aplicar a formatação para as células que contêm os textos específicos
                        cell.fill = header_fill
                        cell.font = header_font


            # Formatação condicional para a coluna PG
            red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            bold_font = Font(bold=True, name="Arial", size=12)
            ws.conditional_formatting.add(f'D{start_row}:D1048576', CellIsRule(operator='equal', formula=['"AB"'], stopIfTrue=True, fill=red_fill, font=bold_font))

            # Itera sobre as células da coluna F (a partir da linha 1 até a última linha com dados)
            for row in range(1, ws.max_row + 1):
                cell = ws[f'F{row}']
                
                # Verificar se o valor da célula é um dos textos desejados
                if cell.value in ['Boletos Abertos', 'Boletos Pagos', 'Total']:
                    # Aplicar o estilo de fonte negrito
                    cell.font = Font(bold=True)
                    cell.font = Font(bold=True, name="Arial", size=12)

            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['E'].width = 20
            
            ws.column_dimensions['F'].width = 50
            ws.column_dimensions['G'].width = 17
            ws.column_dimensions['H'].width = 17
            ws.column_dimensions['I'].width = 60
            
            

        # Salvar a planilha modificada
        wb.save(arquivo_destino)

        formatar_colunas_credito_debito(arquivo_destino,empresa)

        aplicar_bordas_em_todas_as_abas(arquivo_destino,empresa)
        
    except Exception as e:
        infoLogs().info(f"Não foi possível formatar planilha para empresa {empresa}: {e}")
    infoLogs().info("ETAPA FORMATAÇÃO FINAL - FINALIZADA")


       







        
def verifica_pgs_notas_debitos(empresa):
    datas = carregar_datas()
    periodo = ""
    infoLogs().info("ETAPA VERIFICAR STATUS PG NA PLANILHA PRINCIPAL - INICIADA")

    try:
        arquivo_destino = rf'{CAMINHO_DESTINO}\planilha_principal_{empresa}.xlsx'

        planilha_principal = pd.ExcelFile(arquivo_destino)
        abas_principais = planilha_principal.sheet_names

        # Verificar o período e construir o nome da planilha secundária
        if isinstance(datas, tuple):
            periodo = f'{datas[0]}_{datas[1]}'
            planilhaxlsx = f"pagamentos_{periodo}_{empresa}.xlsx"
        elif datas:
            planilhaxlsx = f"pagamentos_{datas}_{empresa}.xlsx"

        planilha_secundaria = pd.read_excel(f"{CAMINHO_NOTAS_PAGAMENTOS}/{planilhaxlsx}")

        # Definir colunas usadas nas planilhas
        coluna_documento = 'NFISCAL'
        coluna_pg = 'PGs'
        coluna_liquidado = 'LIQUIDADO EM'
        coluna_valor = 'valor'
        coluna_credito = 'CREDITO'
        coluna_recbo = 'RECBO'

        # Verificar e tratar duplicatas antes de criar o dicionário
        duplicatas = planilha_secundaria[coluna_documento][planilha_secundaria[coluna_documento].duplicated()]
        if not duplicatas.empty:
            infoLogs().info(f"Existem valores duplicados na coluna '{coluna_documento}':\n{duplicatas}")
            planilha_secundaria = planilha_secundaria.drop_duplicates(subset=[coluna_documento])

        # Função para limpar valores
        def limpar_valor(valor):
            if isinstance(valor, str):
                return valor.strip().replace('\n', '').replace('\r', '')
            else:
                return str(valor).strip()

        # Criar dicionário de dados secundários
        planilha_secundaria[coluna_documento] = planilha_secundaria[coluna_documento].apply(limpar_valor)
        dados_secundarios = planilha_secundaria.set_index(coluna_documento).to_dict(orient='index')

        # Função para atualizar colunas
        def atualizar_colunas(row):
            if row[coluna_pg] in ['PG', 'BX']:
                return row[coluna_pg], row[coluna_credito], row[coluna_recbo]

            if pd.isna(row[coluna_credito]) or row[coluna_credito] == '':
                return row[coluna_pg], row[coluna_credito], row[coluna_recbo]

            nfiscal_principal = limpar_valor(str(row[coluna_documento]))

            if nfiscal_principal in dados_secundarios:
                valor_value = dados_secundarios[nfiscal_principal].get(coluna_valor, row[coluna_credito])
                recbo_value = dados_secundarios[nfiscal_principal].get(coluna_liquidado, row[coluna_recbo])
                return 'PG', valor_value, recbo_value
            else:
                return 'AB', row[coluna_credito], row[coluna_recbo]

        # Iterar pelas abas da planilha principal
        with pd.ExcelWriter(arquivo_destino, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            for aba in abas_principais:
                planilha_atual = pd.read_excel(arquivo_destino, sheet_name=aba)

                if all(col in planilha_atual.columns for col in [coluna_documento, coluna_pg, coluna_credito, coluna_recbo]):
                    planilha_atual[[coluna_pg, coluna_credito, coluna_recbo]] = planilha_atual.apply(
                        atualizar_colunas, axis=1, result_type='expand'
                    )

                planilha_atual.to_excel(writer, sheet_name=aba, index=False)

        infoLogs().info(f"Valores de PG, CREDITO e RECBO atualizados com sucesso em todas as abas | {empresa}")

    except Exception as e:
        infoLogs().error(f"Erro ao processar a planilha pagamentos na data {datas} para empresa {empresa}: \n{e}")

    infoLogs().info("ETAPA VERIFICAR STATUS PG NA PLANILHA PRINCIPAL - FINALIZADA")


    
    


def verifica_pgs_notas_creditos(empresa):
    datas = carregar_datas()
    infoLogs().info("ETAPA VERIFICAR STATUS PG NA PLANILHA PRINCIPAL NOTAS CREDITOS - INICIADA")

    try:
        arquivo_destino = rf'{CAMINHO_DESTINO}\planilha_principal_{empresa}.xlsx'
        planilha_principal = pd.ExcelFile(arquivo_destino)
        abas_principais = planilha_principal.sheet_names

        # Construção do nome da planilha secundária
        if isinstance(datas, tuple):
            periodo = f'{datas[0]}_{datas[1]}'
            planilhaxlsx = f"notas_creditos_{periodo}_pagos_{empresa}.xlsx"
        elif datas:
            planilhaxlsx = f"notas_creditos_{datas}_pagos_{empresa}.xlsx"

        planilha_secundaria = pd.read_excel(f"{CAMINHO_NOTAS_CREDITOS}/{planilhaxlsx}")

        # Definir colunas usadas
        coluna_documento = 'NFISCAL'
        coluna_pg = 'PGs'
        coluna_liquidado = 'LIQUIDADO EM'
        coluna_valor = 'DEBITO'
        coluna_credito = 'CREDITO'
        coluna_recbo = 'RECBO'
        coluna_debito = 'DEBITO'

        # Criar chave única na planilha secundária
        planilha_secundaria['CHAVE_UNICA'] = planilha_secundaria[coluna_documento].astype(str).str.strip() + "_" + planilha_secundaria['NOME'].astype(str).str.strip()
        planilha_secundaria = planilha_secundaria.drop_duplicates(subset='CHAVE_UNICA', keep='first')
        dados_secundarios = planilha_secundaria.set_index('CHAVE_UNICA').to_dict(orient='index')

        def atualizar_colunas(row):
            chave_unica_principal = str(row[coluna_documento]).strip() + "_" + str(row['NOME']).strip()
           

            if row[coluna_pg] in ['PG','BX']:
                return row[coluna_pg], row[coluna_valor], row[coluna_recbo], row[coluna_debito]

            if not pd.isna(row[coluna_credito]) and row[coluna_credito] != '':
                return row[coluna_pg], row[coluna_valor], row[coluna_recbo], row[coluna_debito]

            if pd.isna(row[coluna_valor]) or row[coluna_valor] == '':
                return row[coluna_pg], row[coluna_valor], row[coluna_recbo], row[coluna_debito]

            if chave_unica_principal in dados_secundarios:
                valor_value = dados_secundarios[chave_unica_principal].get(coluna_valor, row[coluna_valor])
                recbo_value = dados_secundarios[chave_unica_principal].get(coluna_liquidado, row[coluna_recbo])
                debito_value = dados_secundarios[chave_unica_principal].get(coluna_debito, row[coluna_debito])

                return 'PG', valor_value, recbo_value, debito_value
            else:
                return 'AB', row[coluna_valor], row[coluna_recbo], row[coluna_debito]

        # Iterar por todas as abas
        with pd.ExcelWriter(arquivo_destino, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            for aba in abas_principais:
                planilha_atual = pd.read_excel(arquivo_destino, sheet_name=aba)

                if all(col in planilha_atual.columns for col in [coluna_documento, coluna_pg, coluna_valor, coluna_recbo, coluna_credito, coluna_debito]):
                    planilha_atual[[coluna_pg, coluna_valor, coluna_recbo, coluna_debito]] = planilha_atual.apply(
                        atualizar_colunas, axis=1, result_type='expand'
                    )

                planilha_atual.to_excel(writer, sheet_name=aba, index=False)

        infoLogs().info(f"Valores de PG, CREDITO, RECBO e DEBITO atualizados com sucesso em todas as abas | {empresa}")

    except Exception as e:
        infoLogs().error(f"Erro ao processar a planilha pagamentos_creditos na data {datas} para empresa {empresa}: \n{e}")

    infoLogs().info("ETAPA VERIFICAR STATUS PG NA PLANILHA PRINCIPAL NOTAS CREDITOS - FINALIZADA")





def ordenar_abas_por_vencto(empresa):
    
    try:
        arquivo = rf'{CAMINHO_DESTINO}\planilha_principal_{empresa}.xlsx'
        
        infoLogs().info("ETAPA DE ORDENAÇÃO DAS ABAS PELA COLUNA VENCTO - INICIADA")
        
        # Carregar o arquivo Excel
        excel = pd.ExcelFile(arquivo)
        abas = excel.sheet_names

        # Criar um novo arquivo Excel com as abas ordenadas
        with pd.ExcelWriter(arquivo, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            for aba in abas:
                # Carregar a aba atual
                df = pd.read_excel(arquivo, sheet_name=aba)

                # Verificar se a coluna VENCTO existe e é a primeira
                if 'VENCTO' in df.columns and df.columns[0] == 'VENCTO':
                    # Converter a coluna VENCTO para o formato datetime, ignorando erros
                    df['VENCTO'] = pd.to_datetime(df['VENCTO'], format='%d/%m/%Y', errors='coerce')

                    # Separar as linhas com datas válidas das inválidas (ou não preenchidas)
                    df_validas = df[df['VENCTO'].notna()].sort_values(by='VENCTO')
                    df_invalidas = df[df['VENCTO'].isna()]

                    # Combinar novamente as linhas, colocando as inválidas no final
                    df_ordenado = pd.concat([df_validas, df_invalidas], ignore_index=True)

                    # Restaurar o formato original da coluna VENCTO
                    df_ordenado['VENCTO'] = df_ordenado['VENCTO'].dt.strftime('%d/%m/%Y').fillna('')

                    # Salvar a aba ordenada
                    df_ordenado.to_excel(writer, sheet_name=aba, index=False)
                else:
                    # Se não houver a coluna VENCTO ou não for a primeira, salvar sem alterações
                    df.to_excel(writer, sheet_name=aba, index=False)
        
        infoLogs().info(f"ETAPA DE ORDENAÇÃO DAS ABAS PELA COLUNA VENCTO PARA EMPRESA {empresa}- FINALIZADA COM SUCESSO")
    except Exception as e:
        infoLogs().error(f"Erro ao ordenar as abas da empresa {empresa}: \n{e}")




def formatar_planilha_nota_debitos(empresa):
    
    

    try:
            # Exemplo de uso para acessar as datas
        datas = carregar_datas()
        periodo = ""

        infoLogs().info("ETAPA FORMATAR PLANILHA NOTAS DEBITOS - INICIADA")
        
        # Nome do arquivo em formato xls
        if isinstance(datas, tuple):
            periodo = f'{datas[0]}_{datas[1]}'
            nome_xls = f"notas_debitos_{periodo}_{empresa}.xls"
        elif datas:
            nome_xls = f"notas_debitos_{datas}_{empresa}.xls"
        
        caminho = CAMINHO_NOTAS_DEBITOS
        
        # Filtrando as colunas específicas e fazendo a leitura do arquivo
        colunas_especificas = ['Vencto', 'Emissão', 'N. Fiscal', 'Cliente', 'Vl. bruto ']
        df = pd.read_excel(f'{caminho}\\{nome_xls}', engine='xlrd', header=7, usecols=colunas_especificas)

        # Tirando todos os valores NaN do arquivo
        df = df.dropna(subset=['Vencto']).copy()

        # Renomeando as colunas para o formato final
        df.rename(columns={
            'Vencto': 'VENCTO',
            'Emissão': 'EMISSAO',
            'N. Fiscal': 'NFISCAL',
            'Cliente': 'NOME',
            'Vl. bruto ': 'CREDITO'
        }, inplace=True)

        # Criar colunas adicionais
        df.insert(1, 'RECBO', '')
        df.insert(3, 'PGs', '')

        # Preencher valores NaN com o valor da célula de cima para as colunas especificadas
        df['NFISCAL'] = df['NFISCAL'].ffill().astype(int)
        df['NOME'] = df['NOME'].ffill()
        df['EMISSAO'] = df['EMISSAO'].ffill()

        # Ordenar as linhas pela coluna VENCTO
        df['VENCTO'] = pd.to_datetime(df['VENCTO'], errors='coerce', dayfirst=True)
        df.sort_values(by='VENCTO', ascending=True, inplace=True, na_position='last')

        # Reaplicar o formato de data na coluna VENCTO
        df['VENCTO'] = df['VENCTO'].dt.strftime('%d/%m/%Y').fillna('')

        # Definindo a nova ordem das colunas
        nova_ordem = ['VENCTO', 'RECBO', 'EMISSAO', 'PGs', 'NFISCAL', 'NOME', 'CREDITO']
        df = df[nova_ordem]

        # Salvar a planilha formatada em xlsx
        if isinstance(datas, tuple):
            periodo = f'{datas[0]}_{datas[1]}'
            nome_novo = f"notas_debitos_{periodo}_{empresa}.xlsx"
        elif datas:
            nome_novo = f"notas_debitos_{datas}_{empresa}.xlsx"

        df.to_excel(f'{caminho}\\{nome_novo}', index=False)
        infoLogs().info(f"Planilha notas débitos formatada e ordenada com sucesso | {empresa}")
    except Exception as e:
        infoLogs().info(f"Erro ao processar arquivo empresa {empresa} notas débitos para formataçaõ: \n{e}")

    infoLogs().info("ETAPA FORMATAR PLANILHA NOTAS DEBITOS - FINALIZADA")




    


def formatar_planilha_notas_credito_aberta(empresa):
    
    try:
            # Exemplo de uso para acessar as datas
        datas = carregar_datas()

        # Nome do arquivo em formato XLS
        if isinstance(datas, tuple):
            periodo = f'{datas[0]}_{datas[1]}'
        elif datas:
            periodo = datas
        else:
            infoLogs().info("Datas não foram fornecidas")
            return

        nome_xls = f"notas_creditos_{periodo}_aberto_{empresa}.xls"
        nome_novo = f"notas_creditos_{periodo}_aberto_{empresa}.xlsx"
        caminho = CAMINHO_NOTAS_CREDITOS

        # Definindo as colunas relevantes para leitura
        colunas = ['Credor', 'Documento', 'Unnamed: 9', 'Unnamed: 10', 'Total']

        num_headers = 11
        tentativas = 0
        while True:
            
            try:
                # Carregar a planilha com as colunas especificadas
                df = pd.read_excel(f'{caminho}\\{nome_xls}', engine='xlrd', header=num_headers,usecols=colunas) #credor na linha 12
                infoLogs().info(f'Num de cabeçalho encontrados {num_headers}')
                break
            except:
                tentativas += 1
                num_headers = num_headers + 1
                # Carregar a planilha com as colunas especificadas
                infoLogs().info(f'Nao encontrado o numero de cabeçalho notas creditos em aberto - Tentando encontrar {tentativas}x')

                if tentativas == 5:
                    infoLogs().info(f'Nao existem planilha notas_credito_aberto para empresa {empresa}')
                    break
                

        
        
        # Filtrar linhas com valores válidos nas colunas 'Unnamed: 10' e 'Total'
        df_filtro = df.dropna(subset=['Unnamed: 10', 'Total']).copy()

        # Preenchendo valores ausentes na coluna 'Credor'
        df_filtro['Credor'] = df_filtro['Credor'].ffill()

        # Renomeando colunas para o formato final
        df_filtro.rename(columns={
            'Credor': 'NOME',
            'Documento': 'NFISCAL',
            'Unnamed: 9': 'EMISSAO',
            'Unnamed: 10': 'VENCTO',
            'Total': 'DEBITO'
        }, inplace=True)

        df_filtro['VENCTO'] = pd.to_datetime(
            df_filtro['VENCTO'], 
            format='%d/%m/%y',  # Substitua pelo formato correto das datas na sua planilha
            errors='coerce'
        ).dt.strftime('%d/%m/%Y')
        
        df_filtro['EMISSAO'] = pd.to_datetime(
            df_filtro['EMISSAO'], 
            format='%d/%m/%y',  # Substitua pelo formato correto das datas na sua planilha
            errors='coerce'
        ).dt.strftime('%d/%m/%Y')


        
        # Adicionando colunas adicionais
        df_filtro.insert(1, 'RECBO', '')
        df_filtro.insert(3, 'PGs', '')
        df_filtro.insert(7, 'CREDITO', '')
        df_filtro.insert(8, 'OBS', '')

        # Reorganizando as colunas na ordem desejada
        nova_ordem = ['VENCTO', 'RECBO', 'EMISSAO', 'PGs', 'NFISCAL', 'NOME', 'CREDITO', 'DEBITO', 'OBS']
        df_filtro = df_filtro[nova_ordem]

       

        # Salvando a planilha formatada no formato XLSX
        df_filtro.to_excel(f'{caminho}\\{nome_novo}', index=False)
        infoLogs().info(f"Planilha notas créditos formatada e ordenada com sucesso | {empresa}")
        infoLogs().info("ETAPA FORMATAR PLANILHA NOTAS CRÉDITOS EM ABERTO - FINALIZADA")

    except Exception as e:
        infoLogs().info(f"Erro ao processar arquivo empresa {empresa} notas créditos em aberto para formatação: \n{e}")




def formatar_planilha_notas_credito(empresa):
  

    try:
            
        datas = carregar_datas()
        periodo = ""

        if isinstance(datas, tuple):
            periodo = f'{datas[0]}_{datas[1]}'
            nome_xls = f"notas_creditos_{periodo}_{empresa}.xls"
            nomeNovo = f"notas_creditos_{periodo}_pagos_{empresa}.xlsx"
        elif datas:
            nome_xls = f"notas_creditos_{datas}_{empresa}.xls"
            nomeNovo = f"notas_creditos_{datas}_pagos_{empresa}.xlsx"

        caminho = CAMINHO_NOTAS_CREDITOS
        
        # Colunas a serem lidas
        colunas = ['Credor', 'Documento', 'Unnamed: 3', 'Unnamed: 9', 'Unnamed: 10', 'Total']

        tentativas = 0
        num_headers = 11
        while True:
            
            try:
                # Carregar a planilha com as colunas especificadas
                df = pd.read_excel(f'{caminho}\\{nome_xls}', engine='xlrd', header=num_headers, usecols=colunas) #credor na linha 12
                infoLogs().info(f'Num de cabeçalho encontrados {num_headers}')
                break
            except:
                tentativas += 1
                num_headers = num_headers + 1
                # Carregar a planilha com as colunas especificadas
                infoLogs().info(f'Nao encontrado o numero de cabeçalho notas creditos em aberto - Tentando encontrar {tentativas}x')

                if tentativas == 5:
                    infoLogs().info(f'Nao existem planilha notas_credito_aberto para empresa {empresa}')
                    break
        

        # Renomear colunas
        df.rename(columns={
            'Unnamed: 3': 'LIQUIDADO EM',
            'Credor': 'NOME',
            'Documento': 'NFISCAL',
            'Unnamed: 9': 'EMISSAO',
            'Unnamed: 10': 'VENCTO',
            'Total': 'DEBITO'
        }, inplace=True)

        # Salvar temporariamente como XLSX
        df.to_excel(f'{caminho}\\{nomeNovo}', index=False)

        # Carregar data da baixa com openpyxl
        wb = load_workbook(f'{caminho}\\{nomeNovo}')
        ws = wb.active
        data_baixa = None
        for row in ws.iter_rows(min_row=1, max_col=4, values_only=False):
            if row[0].value == "Data da baixa":
                data_baixa = row[1].value  # Capturar valor na coluna D
                break

        # Filtrar e preencher dados no DataFrame
        df = pd.read_excel(f'{caminho}\\{nomeNovo}')
        df_filtro = df.dropna(subset=['VENCTO', 'NFISCAL']).copy()

        # df_filtro['VENCTO'] = df_filtro['VENCTO'].strftime('%d/%m/%Y')

        df_filtro.loc[:, 'NOME'] = df_filtro['NOME'].ffill() # Preencher nomes com forward-fill

        df_filtro['LIQUIDADO EM'] = data_baixa  # Adicionar data da baixa

        # Salvar DataFrame final
        df_filtro.to_excel(f'{caminho}\\{nomeNovo}', index=False)
        infoLogs().info(f"Planilha notas créditos formatada e salva com sucesso | {empresa}")

    except Exception as e:
        infoLogs().error(f"Erro ao processar empresa: {empresa} a planilha notas créditos pagos para formatação: \n{str(e)}")


    


def formatar_notas_canceladas(empresa):
    # Exemplo de uso para acessar as datas
    datas = carregar_datas()

    periodo = ""
    infoLogs().info("ETAPA FORMATAR PLANILHA NOTAS DEBITOS CANCELADAS - INICIADA")
    
    try:
        # Nome do arquivo em formato xls
        if isinstance(datas, tuple):
            # Se forem duas datas, imprime as duas
            periodo = f'{datas[0]}_{datas[1]}'
            nome_xls = f"notas_canceladas_{periodo}_{empresa}.xls"
        elif datas:
            nome_xls = f"notas_canceladas_{datas}_{empresa}.xls"
        
        caminho = CAMINHO_NOTAS_CANCELADAS
        
        # Colunas a serem utilizadas
        colunas = ['Emissão', 'Série/Número', 'Cancelamento','Destinatário / Remetente','Motivo']
        
        # Leitura da planilha
        df = pd.read_excel(f'{caminho}\\{nome_xls}', engine='xlrd', header=6,usecols=colunas)
        
       
        # Remover o prefixo 'FL/0' da coluna 'Série/Número'
        df['Série/Número'] = df['Série/Número'].str.replace(r'^FL/0+', '', regex=True)
        
        # Renomear a coluna 'Série/Número' para 'NFISCAL'
        df.rename(columns={'Série/Número': 'NFISCAL'}, inplace=True)
        
        # Remover linhas com valores NaN
        df.dropna(inplace=True)
        
        # Formatar a coluna NFISCAL para número inteiro
        df['NFISCAL'] = df['NFISCAL'].astype(int)
        
        # Caminho e nome do arquivo a ser salvo em formato xlsx
        if isinstance(datas, tuple):
            # Se forem duas datas, imprime as duas
            periodo = f'{datas[0]}_{datas[1]}'
            nome_xlsx = f"notas_canceladas_{periodo}_{empresa}.xlsx"
        elif datas:
            nome_xlsx = f"notas_canceladas_{datas}_{empresa}.xlsx"
        
        caminho_arquivo_xlsx = f"{caminho}\\{nome_xlsx}"
        
        # Salvando o DataFrame formatado em formato XLSX
        df.to_excel(caminho_arquivo_xlsx, index=False)
        
        infoLogs().info(f"Planilha notas_canceladas_formatada | {empresa}")
       
    except Exception as e:
        infoLogs().error(f"Erro ao tentar na empresa {empresa} salvar planilha notas debitos canceladas para formatação: \n{str(e)}")
    
    infoLogs().info("ETAPA FORMATAR PLANILHA NOTAS DEBITOS CANCELADAS - FINALIZADA")
    




def formatarPlanilhaPagamentos(empresa):

    # Exemplo de uso para acessar as datas
    datas = carregar_datas()

    periodo = ""
    try:
        # Carregar a planilha a partir da linha 12
        caminhoPagamentos = CAMINHO_NOTAS_PAGAMENTOS
        
        
        if isinstance(datas, tuple):
            # Se forem duas datas, imprime as duas
            periodo = f'{datas[0]}_{datas[1]}'
            df = pd.read_excel(rf'{caminhoPagamentos}\pagamentos_{periodo}_{empresa}.xls',engine='xlrd', skiprows=16)#alterar quando precissar filtrar qnt de linhas
        elif datas:
            df = pd.read_excel(rf'{caminhoPagamentos}\pagamentos_{datas}_{empresa}.xls',engine='xlrd', skiprows=16)#alterar quando precissar filtrar qnt de linhas


        if 'Unnamed: 27' in df.columns:
            # # # Selecionar as colunas desejadas (ajuste os nomes conforme necessário)
            df = df[['Unnamed: 1','Unnamed: 7', 'Unnamed: 10', 'Unnamed: 11', 'Unnamed: 14', 'Unnamed: 27']]
        
        else:
            df = df[['Unnamed: 1','Unnamed: 7', 'Unnamed: 10', 'Unnamed: 11', 'Unnamed: 14', 'Unnamed: 25']]


        # Renomear as colunas
        df.columns = ['cliente', 'NFISCAL', 'emissao', 'vencimento', 'LIQUIDADO EM', 'valor']



      
        
        df['LIQUIDADO EM'] = pd.to_datetime(
            df['LIQUIDADO EM'],
            format='%d/%m/%y',
            errors='coerce'
        ).dt.strftime('%d/%m/%Y')

        #Tirando todos os valores NaN do arquivo
        df = df.dropna(subset=['NFISCAL'])

        
        # Removendo o texto 'ND 0'
        df['NFISCAL'] = df['NFISCAL'].str.replace('ND 0', '', regex=False)

        # Removendo todos os zeros à esquerda
        df['NFISCAL'] = df['NFISCAL'].str.lstrip('0')
        


        

        if isinstance(datas, tuple):
        # Se forem duas datas, imprime as duas
            periodo = f'{datas[0]}_{datas[1]}'
             #salvar planilha em xlsx
            nomeNovo = f'pagamentos_{periodo}_{empresa}.xlsx'
        elif datas:
            nomeNovo = f'pagamentos_{datas}_{empresa}.xlsx'

        # df['NFISCAL'] = df['NFISCAL'].astype(int)

        # # Salvar a nova planilha
        df.to_excel(rf'{caminhoPagamentos}\{nomeNovo}',index=False)

        infoLogs().info(f'planilha pagamentos formatada com sucesso | {empresa}')
    except Exception as e:
        infoLogs().info(f"Não foi encontrada nenhuma planilha de pagamentos CAR para periodo informado: \n{e}")
    infoLogs().info("ETAPA FORMATAR PLANILHA PAGAMENTOS CAR PG - FINALIZADA")



def formatar_datas_notas(empresa):
    
    caminho_arquivo = rf'{CAMINHO_DESTINO}\planilha_principal_{empresa}.xlsx'  # Substitua pelo caminho do seu arquivo
    wb = load_workbook(caminho_arquivo)
    ws = wb.active  # Seleciona a planilha ativa



    # Definir um estilo para datas abreviadas
    data_style = NamedStyle(name="data_abreviada", number_format="DD/MM/YYYY")


    # Iterar por todas as abas na pasta
    for ws in wb.worksheets:
        # Aplicar o estilo às colunas específicas (exemplo: A e B)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=3):  # Colunas A e B
            for cell in row:
                cell.style = data_style
                
        # Aplicar o formato "Geral" à coluna C
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=5, max_col=5):  # Apenas a coluna C
            for cell in row:
                cell.number_format = "General"  # Define o formato como Geral

    # Salvar as alterações
    wb.save(caminho_arquivo)


def remover_duplicados_chave_unica(colunas_chave, empresa):
    """
    Remove linhas duplicadas em todas as abas de uma planilha com base em uma chave única composta por múltiplas colunas.

    Args:
        colunas_chave (list): Lista das letras das colunas que compõem a chave única (NFISCAL e NOME).
    """
    arquivo_excel = rf'{CAMINHO_DESTINO}\planilha_principal_{empresa}.xlsx'

    # Abre o arquivo Excel
    wb = openpyxl.load_workbook(arquivo_excel)
    
    for aba in wb.sheetnames:  # Itera por todas as abas
        ws = wb[aba]
        infoLogs().info(f"Processando aba: {aba}")

        # Conjunto para armazenar as chaves únicas
        chaves_unicas = set()
        linhas_para_remover = []

        # Percorre as linhas a partir da linha 2 (ignora cabeçalho)
        for linha in range(2, ws.max_row + 1):
            # Construa a chave única concatenando as colunas 'NFISCAL' e 'NOME'
            nfiscal = ws[f'F{linha}'].value  # Coluna NFISCAL (ajuste para o índice correto)
            nome = ws[f'E{linha}'].value  # Coluna NOME (ajuste para o índice correto)
            
            # Garantir que os valores não sejam nulos ou vazios
            if nfiscal and nome:
                chave_unica = f"{str(nome).strip()}_{str(nfiscal).strip()}"  # Chave única sem espaços extras
                
                if chave_unica in chaves_unicas:
                    linhas_para_remover.append(linha)  # Marca a linha para remoção se a chave for duplicada
                else:
                    chaves_unicas.add(chave_unica)

        # Remove as linhas duplicadas, começando pela última
        for linha in reversed(linhas_para_remover):
            ws.delete_rows(linha)

        infoLogs().info(f"Linhas duplicadas removidas da aba: {aba}")

    # Salva as alterações no arquivo Excel
    wb.save(arquivo_excel)
    infoLogs().info(f"Processamento concluído. Linhas duplicadas removidas com base na chave única | {empresa}")


def criar_planilha_principal(empresa):

    arquivo_destino = rf'{CAMINHO_DESTINO}\planilha_principal_{empresa}.xlsx'
    


    # Verificar se o arquivo já existe
    if not os.path.exists(arquivo_destino):
        # Se não existir, cria a planilha
        wb = openpyxl.Workbook()

        # Salvar a planilha criada
        wb.save(arquivo_destino)
        infoLogs().info(f"Arquivo '{arquivo_destino}' criado com sucesso. |{empresa}")
    else:
        infoLogs().info(f"já existe planilha principal para empresa  {empresa}")




def excluir_aba(empresa):
    try:
        
        # Exemplo de uso
        nome_aba = "Sheet"  # Substitua pelo nome da aba que você deseja excluir

        caminho_arquivo = rf'{CAMINHO_DESTINO}\planilha_principal_{empresa}.xlsx'

                # Carregar a planilha
        wb = openpyxl.load_workbook(caminho_arquivo)

        # Verificar se a aba existe
        if nome_aba in wb.sheetnames:
            # Excluir a aba
            del wb[nome_aba]
            wb.save(caminho_arquivo)  # Salvar as mudanças
            infoLogs().info(f"Aba '{nome_aba}' excluída com sucesso | {empresa}")
        else:
            infoLogs().info(f"Aba '{nome_aba}' não existe. Nenhuma ação foi realizada | {empresa}")

    except Exception as e:
        infoLogs().info(f"Erro ao tentar excluir a aba empresa {empresa}: {e}")


def reordenar_credito_debito(empresa):


    try:
        # Carregar o arquivo Excel com várias abas
        file_path = rf'{CAMINHO_DESTINO}\planilha_principal_{empresa}.xlsx'  # Caminho da planilha
        
        output_path = rf'{CAMINHO_DESTINO}\planilha_principal_{empresa}.xlsx'  # Caminho da planilha

        # Carregar todas as abas do Excel em um dicionário de DataFrames
        sheets = pd.read_excel(file_path, sheet_name=None)

        # Inicializar um dicionário para armazenar os DataFrames organizados
        organized_sheets = {}

        # Iterar sobre cada aba
        for sheet_name, df in sheets.items():
            # Verificar se as colunas necessárias existem na aba atual
            if {'VENCTO', 'CREDITO', 'DEBITO'}.issubset(df.columns):
                # Criar coluna auxiliar para ordenar
                df['CREDITO_PRESENT'] = df['CREDITO'].notnull().astype(int)
                
                # Ordenar por VENCTO e garantir CREDITO antes de DEBITO
                organized_df = df.sort_values(by=['VENCTO', 'CREDITO_PRESENT'], ascending=[True, False])
                
                # Remover a coluna auxiliar
                organized_df = organized_df.drop(columns=['CREDITO_PRESENT'])
                
                # Salvar o DataFrame organizado no dicionário
                organized_sheets[sheet_name] = organized_df
            else:
                # Se as colunas necessárias não existirem, manter a aba original
                organized_sheets[sheet_name] = df

        # Salvar todas as abas organizadas em um novo arquivo Excel
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            for sheet_name, organized_df in organized_sheets.items():
                organized_df.to_excel(writer, sheet_name=sheet_name, index=False)

        infoLogs().info(f"Planilha empresa {empresa} ordenada CREDITO e DEBITO")
    except Exception as e:
        infoLogs().info(f"Erro empresa {empresa} para ordenar CREDITO e DEBITO | {e}")






# Função para extrair o mês e o ano do nome da aba
def extrair_mes_ano(aba):
    try:
        # Dividir o nome da aba e obter o mês e o ano
        partes = aba.split('-')
        mes = partes[0]
        ano = int(partes[1])
        
        # Converter o nome do mês para um número
        meses = {
            'Janeiro': 1, 'Fevereiro': 2, 'Marco': 3, 'Abril': 4,
            'Maio': 5, 'Junho': 6, 'Julho': 7, 'Agosto': 8,
            'Setembro': 9, 'Outubro': 10, 'Novembro': 11, 'Dezembro': 12
        }
        mes_num = meses[mes]
        return (ano, mes_num)
    except Exception as e:
        infoLogs().info(f'Erro ao extrair mês e ano da aba {aba} |')
        return (0, 0)  # Retornar um valor padrão em caso de erro

def reordenar_por_ano_abas(empresa):
    # Definir o caminho do arquivo de entrada (e saída, que é o mesmo)
    file_path = rf'{CAMINHO_DESTINO}\planilha_principal_{empresa}.xlsx'

    # Carregar o arquivo Excel usando openpyxl
    wb = load_workbook(file_path)

    # Obter os nomes das abas
    abas = wb.sheetnames

    # Ordenar as abas pelo ano e mês
    try:
        abas_ordenadas = sorted(abas, key=extrair_mes_ano)
        # Reorganizar as abas no arquivo Excel
        # Para isso, vamos mover as abas para a ordem desejada
        for aba in abas_ordenadas:
            # Move a aba para a primeira posição (ou para a posição desejada)
            wb[aba].sheet_state = 'visible'  # Garantir que a aba esteja visível, se necessário
            wb._sheets.sort(key=lambda sheet: abas_ordenadas.index(sheet.title))

        # Salvar o arquivo Excel com as abas reordenadas
        wb.save(file_path)

        infoLogs().info(f"Abas ordenadas por ano e mês da empresa '{empresa}'.")
    except Exception as e:
        infoLogs().info(f'Erro ao reordenar dados {empresa} - {e}')


    







def formatacao_planilha_final():
    
    for i in EMPRESAS:
            
        match i:
            case 1:
                empresa = 'lokan'
            case 2:
                empresa = 'lokanestrutura'  
            case 3:
                empresa ='lokanfacil'
            case 4:
                empresa = 'lokanlondrina'


        
        formatarPlanilhaPagamentos(empresa)
        
        formatar_planilha_notas_credito(empresa)
        
        formatar_planilha_nota_debitos(empresa)
        
        formatar_notas_canceladas(empresa)
        
        formatar_planilha_notas_credito_aberta(empresa)

        criar_planilha_principal(empresa)

        limpar_boletos_e_excluir_cabeçalhos(empresa)

        

        copia_notas_debitos(empresa)

        copia_notas_creditos_em_aberto(empresa)
        
        copia_notas_canceladas(empresa)

        ordenar_abas_por_vencto(empresa)

        reordenar_credito_debito(empresa)

        reordenar_por_ano_abas(empresa)
        

        verifica_pgs_notas_debitos(empresa)

        verifica_pgs_notas_creditos(empresa)
        
        # Uso da função
        colunas_chave=['F', 'E']
        remover_duplicados_chave_unica(colunas_chave,empresa)  # Colunas: NFISCAL (E), NOME (F),))

        reordenar_credito_debito(empresa)

        separar_datas(empresa)

        formatar_datas_notas(empresa)

        formatacaoFinal(empresa)
        
        excluir_aba(empresa)


